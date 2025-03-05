import openpyxl

def get_cell_data(filename, sheetname, row, col):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.cell(row, col).value

def get_all_data(filename, sheet):
    mainList = []
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet]
    row = sheet.max_row
    col = sheet.max_column
    for i in range(2, row+1):
        list1 = []
        for j in range(1, col+1):
            list1.append(sheet.cell(i,j).value)
        mainList.append(list1)
    return mainList



